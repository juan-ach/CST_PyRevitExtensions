# -*- coding: utf-8 -*-
"""
Helper externo: lee la hoja "3. ELEM. PRINCIPALES I.F." del fichero Excel
indicado como argumento y devuelve un JSON {clave_colC: valor_colG_ajustado} por stdout.

Regla de ajuste:
- Si el valor de la columna D en la fila es 1, se usa G tal cual.
- Si el valor de la columna D es distinto de 1, se usa G / D.

Uso:
    python read_excel_helper.py "<ruta_excel>" "<nombre_hoja>"

Requiere openpyxl instalado en el Python del sistema:
    pip install openpyxl
"""
import sys
import os
import json


def _to_float(value):
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        try:
            return float(str(value).replace(",", ".").strip())
        except (ValueError, TypeError):
            return None


def main():
    if len(sys.argv) < 3:
        error = {"error": "Uso: read_excel_helper.py <excel_path> <sheet_name>"}
        print(json.dumps(error))
        sys.exit(1)

    excel_path = sys.argv[1]
    sheet_name = sys.argv[2]

    if not os.path.isfile(excel_path):
        print(json.dumps({"error": "Fichero no encontrado: {}".format(excel_path)}))
        sys.exit(1)

    try:
        import openpyxl
    except ImportError:
        # Auto-instalar en el Python del sistema si no esta
        import subprocess
        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", "openpyxl", "--quiet"]
        )
        import openpyxl

    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    except Exception as e:
        print(json.dumps({"error": "Error al abrir Excel: {}".format(str(e))}))
        sys.exit(1)

    if sheet_name not in wb.sheetnames:
        print(json.dumps({
            "error": "Hoja '{}' no encontrada. Disponibles: {}".format(
                sheet_name, ", ".join(wb.sheetnames)
            )
        }))
        sys.exit(1)

    ws = wb[sheet_name]
    IDX_C = 2  # Columna C (0-based)
    IDX_D = 3  # Columna D (0-based)
    IDX_G = 6  # Columna G (0-based)
    MAX_ROW = 300

    resultado = {}
    none_g = []
    invalid_d = []
    zero_d = []

    for row in ws.iter_rows(min_row=2, max_row=MAX_ROW, values_only=True):
        if len(row) <= IDX_G:
            continue

        clave = row[IDX_C]
        valor_d = row[IDX_D] if len(row) > IDX_D else None
        valor_g = row[IDX_G]

        if clave is None:
            continue

        clave_str = str(clave).strip()
        if not clave_str:
            continue

        # Si la clave ya fue encontrada antes, la omitimos (nos quedamos con el primer valor)
        if clave_str in resultado:
            continue

        valor_g_num = _to_float(valor_g)
        if valor_g_num is None:
            none_g.append(clave_str)
            resultado[clave_str] = None
            continue

        divisor_d = _to_float(valor_d)
        if divisor_d is None:
            invalid_d.append(clave_str)
            resultado[clave_str] = None
            continue

        if divisor_d == 0:
            zero_d.append(clave_str)
            resultado[clave_str] = None
            continue

        if divisor_d == 1:
            resultado[clave_str] = valor_g_num
        else:
            resultado[clave_str] = valor_g_num / divisor_d

    wb.close()

    output = {
        "data": resultado,
        "none_g": none_g,
        "invalid_d": invalid_d,
        "zero_d": zero_d,
        "sheets": wb.sheetnames,
        "max_row": MAX_ROW
    }
    print(json.dumps(output, ensure_ascii=True))


if __name__ == "__main__":
    main()
